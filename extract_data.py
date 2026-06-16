#!/usr/bin/env python
"""
Extract weekly report data from Excel to JSON.
Usage: python extract_data.py <excel_file.xlsx> [output.json]

Reads the Anta retail weekly report Excel format and extracts
all data tables into extracted_data.json for build_dashboard.py
and generate_reports.py.

Excel structure (based on actual 周报打印.xlsx):
 - Sheet: '周报' — main KPI, daily, category, series, TOP, etc.
 - Sheet: '人员销售' — staff sales performance
 - Sheet: '产品销售' — seasonal product breakdown
"""
import json, sys, os, decimal


def cv_safe(ws, row, col):
    """Get cell value safely, handling errors and special values."""
    cell = ws.cell(row=row, column=col)
    if cell is None or cell.value is None:
        return None
    v = cell.value
    if isinstance(v, str) and v in ('#DIV/0!', '/', 'DIV/0', '—', '-'):
        return None
    if isinstance(v, decimal.Decimal):
        return float(v)
    if isinstance(v, (int, float)):
        return v
    # Try numeric conversion for number-like strings
    if isinstance(v, str):
        v_stripped = v.strip().replace(',', '').replace('￥', '').replace('¥', '')
        try:
            return float(v_stripped) if '.' in v_stripped else int(v_stripped)
        except (ValueError, TypeError):
            return v  # return as string (label/header)
    return v


def scan_labeled_rows(ws, start_row, end_row, label_col=1, data_cols=None):
    """
    Scan rows from start_row to end_row, extracting rows that have a non-empty
    label in label_col. Returns dict {str(row_num): {'label':..., 'data':{col:val}}}.
    """
    if data_cols is None:
        data_cols = list(range(1, 42))
    result = {}
    for r in range(start_row, end_row + 1):
        label = cv_safe(ws, r, label_col)
        if label is not None:
            label_str = str(label).strip()
            if label_str:
                entry = {'label': label_str, 'data': {}}
                for c in data_cols:
                    v = cv_safe(ws, r, c)
                    if v is not None:
                        entry['data'][str(c)] = v
                result[str(r)] = entry
    return result


def find_row(ws, keyword, col=1, max_row=250):
    """Find row number containing keyword in given column."""
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=col).value
        if cell and keyword in str(cell):
            return row
    return 0


def extract(excel_path):
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ── Identify sheets ──
    ws_main = wb['周报'] if '周报' in wb.sheetnames else None
    if ws_main is None:
        for name, ws in zip(wb.sheetnames, [wb[s] for s in wb.sheetnames]):
            if find_row(ws, 'KPI', 1, 20):
                ws_main = ws
                break
    if ws_main is None:
        raise ValueError("Cannot find main data sheet")

    ws_seas = wb['产品销售'] if '产品销售' in wb.sheetnames else None
    if ws_seas is None:
        for name, ws in zip(wb.sheetnames, [wb[s] for s in wb.sheetnames]):
            if find_row(ws, '产品季', 4):
                ws_seas = ws
                break

    ws_member = wb['人员销售'] if '人员销售' in wb.sheetnames else None
    if ws_member is None:
        for name, ws in zip(wb.sheetnames, [wb[s] for s in wb.sheetnames]):
            if find_row(ws, '工号', 1):
                ws_member = ws
                break

    result = {}

    # ════════════════════════════════════════════
    # 1) r7_kpi: main KPI row
    # ════════════════════════════════════════════
    kpi_row = find_row(ws_main, 'KPI') or 7
    data_kpi_row = kpi_row + 3 if str(cv_safe(ws_main, kpi_row, 1) or '') == 'KPI' else kpi_row
    kpi_cols = [4, 5, 6, 7, 8, 10, 12, 14, 17, 20, 22, 24, 26, 28, 30, 32, 33, 35, 36, 37, 39]
    r7_kpi = {}
    for c in kpi_cols:
        r7_kpi[str(c)] = cv_safe(ws_main, data_kpi_row, c)
    result['r7_kpi'] = r7_kpi

    # ════════════════════════════════════════════
    # 2) r15_price: price metrics row
    # ════════════════════════════════════════════
    price_row = find_row(ws_main, '件单价') or data_kpi_row + 8
    # Find the next '周合计' row after the header
    for dr in range(price_row, price_row + 10):
        v = cv_safe(ws_main, dr, 1)
        if v and str(v).strip() in ('周合计', '周'):
            price_row = dr
            break
    price_cols = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32]
    r15_price = {}
    for c in price_cols:
        r15_price[str(c)] = cv_safe(ws_main, price_row, c)
    result['r15_price'] = r15_price

    # ════════════════════════════════════════════
    # 3) Daily breakdown (rows 21-30)
    # ════════════════════════════════════════════
    daily_rows = [21, 22, 23, 24, 25, 26, 27, 28, 29, 30]
    day_cols = [4, 6, 8, 10, 12, 14, 16]
    day_col_keys = ['4', '6', '8', '10', '12', '14', '16']
    daily = {}
    for row_num in daily_rows:
        dk = str(row_num)
        daily[dk] = {'data': {}}
        for i, dc in enumerate(day_cols):
            v = cv_safe(ws_main, row_num, dc)
            daily[dk]['data'][day_col_keys[i]] = v if v is not None else 0
    result['daily'] = daily

    # ════════════════════════════════════════════
    # 4-11) All labeled row sections
    # ════════════════════════════════════════════
    cate_cols = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34, 36, 38, 40]
    series_cols = cate_cols
    top_cols = list(range(1, 35))
    disc_cols = list(range(1, 25))

    result['category'] = scan_labeled_rows(ws_main, 35, 57, data_cols=cate_cols)
    result['clothing_series'] = scan_labeled_rows(ws_main, 62, 74, data_cols=series_cols)
    result['shoe_series'] = scan_labeled_rows(ws_main, 75, 88, data_cols=series_cols)
    result['sub_ps'] = scan_labeled_rows(ws_main, 90, 109, data_cols=series_cols)
    result['accessory_sub'] = scan_labeled_rows(ws_main, 111, 122, data_cols=series_cols)
    result['top_goods'] = scan_labeled_rows(ws_main, 123, 150, data_cols=top_cols)
    result['discount_bracket'] = scan_labeled_rows(ws_main, 155, 195, data_cols=disc_cols)
    result['discount_range'] = result['discount_bracket']
    result['order_structure'] = scan_labeled_rows(ws_main, 190, 210, data_cols=list(range(1, 25)))

    # ════════════════════════════════════════════
    # 12) Seasonal data (产品销售 sheet)
    # ════════════════════════════════════════════
    if ws_seas:
        seas_cols = list(range(1, min(ws_seas.max_column + 1, 53)))
        result['seasonal'] = scan_labeled_rows(ws_seas, 1, ws_seas.max_row,
                                               label_col=1, data_cols=seas_cols)
    else:
        result['seasonal'] = {}

    # ════════════════════════════════════════════
    # 13) Member / staff data (人员销售 sheet)
    # ════════════════════════════════════════════
    if ws_member:
        member_cols = list(range(1, min(ws_member.max_column + 1, 37)))
        result['member'] = scan_labeled_rows(ws_member, 5, ws_member.max_row,
                                             label_col=1, data_cols=member_cols)
    else:
        result['member'] = {}

    # ════════════════════════════════════════════
    # 14) Report config — auto-detect Wxx version
    # ════════════════════════════════════════════
    import re
    period_detected = 'W24'
    for r in range(1, 10):
        v = cv_safe(ws_main, r, 23)
        if v and isinstance(v, str) and 'W' in v and '周' in v:
            m = re.search(r'(W\d+)', v)
            if m:
                period_detected = m.group(1)
                break
    result['report_config'] = {'_version': period_detected}

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_data.py <excel_file.xlsx> [output.json]")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'extracted_data.json'

    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found")
        sys.exit(1)

    print(f"Extracting from {excel_path}...")
    data = extract(excel_path)

    class CustomEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, decimal.Decimal):
                return float(obj)
            return super().default(obj)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, cls=CustomEncoder)

    for sec in ['r7_kpi', 'r15_price', 'daily', 'category', 'clothing_series',
                'shoe_series', 'sub_ps', 'accessory_sub', 'top_goods',
                'discount_bracket', 'order_structure', 'seasonal', 'member']:
        d = data.get(sec, {})
        if isinstance(d, dict) and d:
            first = next(iter(d.values()), {})
            if isinstance(first, dict) and 'data' in first:
                print(f"  {sec}: {len(d)} rows")
            else:
                print(f"  {sec}: {len(d)} keys")

    period = data.get('report_config', {}).get('_version', 'unknown')
    print(f"  Detected period: {period}")
    print(f"Data extracted to {output_path}")


if __name__ == '__main__':
    main()
